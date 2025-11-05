# app/routes.py
from flask import Blueprint, request, jsonify, Response, stream_with_context, current_app, send_file
from app.models import Dispositivo, EstadoLog
from app.db import db
from app.models import (
    Dispositivo,
    EstadoLog,
    Habitacion,
    User,
    UserProfileExtra,   # perfil extra (avatar + preferencia de tema)
    SecurityCode,       # códigos de seguridad (cambio contraseña y forgot/registro)
    AccionLog           # <-- NUEVO: para auditoría/eventos de negocio
)
from app.sse import subscribe, unsubscribe, publish as sse_publish
import json, time, requests
from queue import Empty
from datetime import datetime, timezone
from app.iotelligence.core import dispatch_measure

import os
import smtplib
import ssl
import secrets
from email.message import EmailMessage
from email.utils import make_msgid
from datetime import datetime, timedelta
from queue import Empty
from secrets import randbelow
from flask_jwt_extended import create_access_token, jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from app.iotelligence.dev_kinds import infer_kind, infer_capability

# NUEVO: utilidades para exportar Excel
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

bp = Blueprint('routes', __name__)

def _disp_payload(d: Dispositivo) -> dict:
    cfg = d.configuracion or {}
    kind = infer_kind(d.serial_number or "", cfg)
    capability = infer_capability(kind, cfg)
    return {
        "id": d.id,
        "serial_number": d.serial_number,
        "nombre": d.nombre,
        "tipo": d.tipo,
        "modelo": d.modelo,
        "descripcion": d.descripcion,
        "estado": d.estado,
        "parametros": d.parametros,
        "configuracion": cfg,
        "reclamado": d.reclamado,
        # NUEVO:
        "kind": kind,
        "capability": capability,
    }

# -------------------------------------------------------------------
# NUEVO: Helper para registrar acciones (auditoría AccionLog)
# -------------------------------------------------------------------
def _actor_tag() -> str:
    # Si hay usuario autenticado devolvemos 'user', si no 'system'
    try:
        uid = get_jwt_identity()
        return "user" if uid is not None else "system"
    except Exception:
        return "system"

def _log_action(dispositivo_id: int, evento: str, detalle: dict | None = None, actor: str | None = None):
    try:
        a = AccionLog(
            dispositivo_id=dispositivo_id,
            evento=evento,
            detalle=(detalle or {}),
            actor=(actor or _actor_tag())
        )
        db.session.add(a)
        # No hacemos commit aquí; se comitea en la transacción del caller
    except Exception as e:
        current_app.logger.warning(f"[accion_log] No se pudo registrar acción {evento} para {dispositivo_id}: {e}")

@bp.route('/dispositivos', methods=['POST'])
@jwt_required(optional=True)
def agregar_dispositivo():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Datos JSON inválidos o faltantes"}), 400

        nuevo = Dispositivo(
            serial_number=data.get('serial_number'),
            nombre=data.get('nombre'),
            tipo=data.get('tipo'),
            modelo=data.get('modelo', ''),
            descripcion=data.get('descripcion', ''),
            estado=data.get('estado', 'desconocido'),
            parametros={},
            configuracion=data.get('configuracion', {}),
            reclamado=data.get('reclamado', False)  # <-- Nuevo campo opcional
        )
        db.session.add(nuevo)
        db.session.commit()
        return jsonify({"mensaje": "Dispositivo agregado", "id": nuevo.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error al agregar dispositivo", "detalle": str(e)}), 500


def _safe_get(obj, *names, default=None):
    """Devuelve el primer atributo existente en 'names' dentro de obj."""
    for n in names:
        if hasattr(obj, n):
            return getattr(obj, n)
    return default


def _safe_set(obj, value, *names):
    """Setea el primer atributo existente en 'names' con 'value'."""
    for n in names:
        if hasattr(obj, n):
            setattr(obj, n, value)
            return True
    return False


def _coerce_bool(val):
    if isinstance(val, bool):
        return val
    if val is None:
        return None
    s = str(val).strip().lower()
    return s in ("true", "1", "t", "yes", "y")


# ---- Dispositivo: getters/setters tolerantes a nombres ----

def _dev_get_id(d):
    return _safe_get(d, "id")


def _dev_get_serial(d):
    # Acepta múltiples nombres de atributo; preferimos 'serial_number' si existe
    return _safe_get(d, "serial_number", "serial", "serialNumber", "uuid", "mac", default=None)


def _dev_set_serial(d, v):
    return _safe_set(d, v, "serial_number", "serial", "serialNumber", "uuid", "mac")


def _dev_get_tipo(d):
    return _safe_get(d, "tipo", "type", "device_type", "deviceType", default=None)


def _dev_set_tipo(d, v):
    return _safe_set(d, v, "tipo", "type", "device_type", "deviceType")


def _dev_get_nombre(d):
    return _safe_get(d, "nombre", "name", "label", default=None)


def _dev_set_nombre(d, v):
    return _safe_set(d, v, "nombre", "name", "label")


def _dev_get_reclamado(d):
    val = _safe_get(d, "reclamado", "claimed", default=False)
    return bool(val)


def _dev_set_reclamado(d, v: bool):
    return _safe_set(d, bool(v), "reclamado", "claimed")


def _dev_get_habitacion_id(d):
    return _safe_get(d, "habitacion_id", "room_id", "habitacionId", "roomId", default=None)


def _dev_set_habitacion_id(d, v):
    return _safe_set(d, v, "habitacion_id", "room_id", "habitacionId", "roomId")


def _dev_get_estado(d):
    return _safe_get(d, "estado", "status", "state", default=None)


def _dev_set_estado(d, v):
    return _safe_set(d, v, "estado", "status", "state")


def _device_to_dict(d):
    """
    Serializa SIEMPRE con la clave 'serial_number' (snake_case) para
    alinearse con el cliente Android (Moshi).
    """
    return {
        "id":            _dev_get_id(d),
        "serial_number": _dev_get_serial(d),
        "nombre":        _dev_get_nombre(d),
        "tipo":          _dev_get_tipo(d),
        "reclamado":     _dev_get_reclamado(d),
        "habitacion_id": _dev_get_habitacion_id(d),
        "estado":        _dev_get_estado(d),
    }


def _device_full_payload(d):
    """
    Payload completo y consistente para todas las respuestas 'de dispositivo'.
    """
    base = _device_to_dict(d)
    base.update({
        "modelo": getattr(d, "modelo", "") or "",
        "descripcion": getattr(d, "descripcion", "") or "",
        "parametros": (getattr(d, "parametros", {}) or {}),
        "configuracion": (getattr(d, "configuracion", {}) or {}),
    })
    return base


def _room_to_dict(room: Habitacion):
    payload = {"id": room.id, "nombre": room.nombre}
    icon = _safe_get(room, "icon", "icono", "icon_name", "iconName", default=None)
    if icon is not None:
        payload["icon"] = icon
    return payload


def _set_room_icon_if_present(room: Habitacion, value: str | None):
    if not value:
        return
    _safe_set(room, value, "icon", "icono", "icon_name", "iconName")


@bp.route('/dispositivos/<int:id>', methods=['PUT'])
@jwt_required(optional=True)
def actualizar_dispositivo(id):
    """
    Actualiza un dispositivo respetando:
      - Merge de 'configuracion' y de 'parametros' (no se pisan claves no enviadas).
      - Acepta claves planas (pos/position, speed/velocidad, set_temp, minutes_left, watering_end_epoch, encendido).
      - 'estado' se respeta tal cual venga (on/off, open/closed, locked/unlocked, watering, cleaning, etc.).
      - Si llega 'estado' on/off/activo/inactivo y no llega 'encendido', derivamos 'encendido'.
    Devuelve SIEMPRE el dispositivo completo para que el frontend pueda deserializarlo como Dispositivo.
    """
    try:
        data = request.get_json() or {}
        dispositivo = Dispositivo.query.get_or_404(id)

        old_name = dispositivo.nombre  # para log de renombrado

        # --- CONFIGURACIÓN ---
        cfg_actual = dispositivo.configuracion or {}
        cfg_payload = data.get("configuracion") or {}
        cfg_merged = {**cfg_actual, **cfg_payload}

        # ✅ encendido puede venir arriba o dentro de configuracion
        if "encendido" in data and data["encendido"] is not None:
            cfg_merged["encendido"] = bool(data["encendido"])

        # ✅ Soporte capabilities/capability (plantillas nuevas)
        if "capability" not in cfg_merged and "capabilities" in cfg_merged:
            caps = cfg_merged.pop("capabilities")
            if isinstance(caps, list) and caps:
                cfg_merged["capability"] = caps[0]
            else:
                cfg_merged["capability"] = caps  # string o None tal cual

        # ✅ Asegura que los canales existan como dict si el cliente manda null
        for k in ("horarios", "horarios_pos", "horarios_speed",
                  "horarios_lock", "horarios_riego", "horarios_temp"):
            if k in cfg_merged and cfg_merged[k] is None:
                cfg_merged[k] = {}

        # (opcional) Validación/sanitización ligera de los mapas de horarios
        def _is_event_list(v):
            # espera lista de pares ["HH:MM", valor]
            if not isinstance(v, list): 
                return False
            return all(isinstance(x, (list, tuple)) and len(x) == 2 for x in v)

        def _sanitize_schedule_map(m):
            if not isinstance(m, dict):
                return m
            out = {}
            for day, lst in m.items():
                if _is_event_list(lst):
                    out[str(day).lower()] = lst
            return out

        for ch in ("horarios", "horarios_pos", "horarios_speed",
                   "horarios_lock", "horarios_riego", "horarios_temp"):
            if ch in cfg_merged and isinstance(cfg_merged[ch], dict):
                cfg_merged[ch] = _sanitize_schedule_map(cfg_merged[ch])

        # --- PARAMETROS ---
        par_actual = dispositivo.parametros or {}
        par_payload = data.get("parametros") or {}

        # Extraer campos planos compatibles y meterlos en parámetros
        flat_params = {}
        # pos/position (0..100)
        if "pos" in data and data["pos"] is not None:
            try: flat_params["pos"] = int(data["pos"])
            except: pass
        if "position" in data and data["position"] is not None:
            try: flat_params["position"] = int(data["position"])
            except: pass
        # speed/velocidad
        if "speed" in data and data["speed"] is not None:
            try:
                flat_params["speed"] = int(data["speed"])
                flat_params["velocidad"] = int(data["speed"])  # espejo común
            except:
                pass
        if "velocidad" in data and data["velocidad"] is not None:
            try:
                flat_params["velocidad"] = int(data["velocidad"])
                flat_params["speed"] = int(data["velocidad"])
            except:
                pass
        # set_temp
        if "set_temp" in data and data["set_temp"] is not None:
            try: flat_params["set_temp"] = int(data["set_temp"])
            except: pass
        # minutos restantes
        if "minutes_left" in data and data["minutes_left"] is not None:
            try: flat_params["minutes_left"] = int(data["minutes_left"])
            except: pass
        # epoch de fin de riego
        if "watering_end_epoch" in data and data["watering_end_epoch"] is not None:
            try: flat_params["watering_end_epoch"] = int(data["watering_end_epoch"])
            except: pass

        parametros_merged = {**par_actual, **flat_params, **par_payload}

        # --- ESTADO ---
        estado_in = data.get("estado", None)
        if isinstance(estado_in, str):
            estado_in = estado_in.strip()
            # mantenemos tal cual (sin normalizar a "activo/inactivo")
            if estado_in != "":
                dispositivo.estado = estado_in

        # Derivar encendido si no vino y el estado es un on/off típico
        if "encendido" not in cfg_merged:
            if isinstance(estado_in, str):
                low = estado_in.lower()
                if low in ("on", "activo", "active", "true", "1"):
                    cfg_merged["encendido"] = True
                elif low in ("off", "inactivo", "inactive", "false", "0"):
                    cfg_merged["encendido"] = False
                # Otros estados (open/closed/locked/unlocked/watering/cleaning…) no definen encendido.

        # --- CAMPOS EXTRA (opcionales) ---
        renamed = False
        if "nombre" in data and data["nombre"] != dispositivo.nombre:
            renamed = True
        if "nombre" in data:
            dispositivo.nombre = data["nombre"]
        if "tipo" in data:
            dispositivo.tipo = data["tipo"]
        if "modelo" in data:
            dispositivo.modelo = data["modelo"]
        if "descripcion" in data:
            dispositivo.descripcion = data["descripcion"]

        # Aplicar merges finales en el modelo
        dispositivo.parametros = parametros_merged
        dispositivo.configuracion = cfg_merged

        # Logs: estado histórico
        log = EstadoLog(
            dispositivo_id=dispositivo.id,
            estado=dispositivo.estado,
            parametros=dispositivo.parametros or {}
        )
        db.session.add(log)

        # Logs: acciones/configuración/renombrado
        if renamed:
            _log_action(dispositivo.id, "renamed", {"old": old_name, "new": dispositivo.nombre})
        # Registrar el payload exacto que llegó (útil para auditoría/config)
        if (cfg_payload or par_payload or flat_params or estado_in is not None):
            _log_action(
                dispositivo.id,
                "config_changed",
                {"payload": data}
            )

        db.session.commit()

        # Disparar regla por cambio de config/estado
        now = datetime.now(timezone.utc)
        try:
            dispatch_measure(dispositivo, None, None, ts=now)
        except Exception as e:
            current_app.logger.warning(f"[dispatch_measure] {e}")

        # SSE
        try:
            sse_publish({
                "event": "device_update",
                "id": dispositivo.id,
                "serial_number": dispositivo.serial_number,
                "nombre": dispositivo.nombre,
                "tipo": dispositivo.tipo,
                "modelo": dispositivo.modelo,
                "descripcion": dispositivo.descripcion,
                "estado": dispositivo.estado,
                "parametros": dispositivo.parametros,
                "configuracion": dispositivo.configuracion,
                "reclamado": dispositivo.reclamado
            })
        except Exception as e:
            current_app.logger.warning(f"[sse] publish error: {e}")

        # ✅ DEVOLVER OBJETO COMPLETO
        return jsonify(_device_full_payload(dispositivo)), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error al actualizar dispositivo", "detalle": str(e)}), 500

@bp.route('/dispositivos/<int:id>', methods=['GET'])
@jwt_required(optional=True)
def obtener_dispositivo(id):
    try:
        dispositivo = Dispositivo.query.get_or_404(id)
        return jsonify(_disp_payload(dispositivo))
    except Exception as e:
        return jsonify({"error": "Error al obtener dispositivo", "detalle": str(e)}), 500


@bp.route('/dispositivos', methods=['GET'])
@jwt_required(optional=True)
def obtener_todos():
    try:
        q = Dispositivo.query

        # Filtros opcionales:
        estado = request.args.get('estado')
        if estado:
            q = q.filter(Dispositivo.estado == estado)

        reclamado = request.args.get('reclamado')
        if reclamado in ('true', 'false'):
            q = q.filter(Dispositivo.reclamado == (reclamado == 'true'))

        unassigned = request.args.get('unassigned')
        if unassigned in ('true', 'false'):
            if unassigned == 'true':
                q = q.filter(Dispositivo.habitacion_id.is_(None))
            else:
                q = q.filter(Dispositivo.habitacion_id.is_not(None))

        dispositivos = q.order_by(Dispositivo.id.desc()).all()
        return jsonify([
            {
                **_device_full_payload(d),
                "kind": infer_kind(d.serial_number or "", d.configuracion or {}),
                "capability": infer_capability(
                    infer_kind(d.serial_number or "", d.configuracion or {}),
                    d.configuracion or {}
                )
            }
            for d in dispositivos
        ])
    except Exception as e:
        return jsonify({"error": "Error al obtener dispositivos", "detalle": str(e)}), 500


@bp.route('/dispositivos/<int:id>/logs', methods=['GET'])
@jwt_required(optional=True)
def obtener_logs(id):
    try:
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 50)), 200)

        q = EstadoLog.query.filter_by(dispositivo_id=id).order_by(EstadoLog.timestamp.desc())
        items = q.paginate(page=page, per_page=per_page, error_out=False)

        return jsonify({
            "page": page,
            "per_page": per_page,
            "total": items.total,
            "pages": items.pages,
            "data": [
                {
                    "estado": log.estado,
                    "parametros": log.parametros,
                    "timestamp": log.timestamp.isoformat()
                } for log in items.items
            ]
        })
    except Exception as e:
        return jsonify({"error": "Error al obtener logs", "detalle": str(e)}), 500


@bp.route('/dispositivos/reclamar', methods=['POST'])
@jwt_required(optional=True)
def reclamar_dispositivo():
    try:
        data = request.get_json() or {}
        serial = data.get('serial_number')
        if not serial:
            return jsonify({"error":"Debe proporcionar serial_number"}), 400

        disp = Dispositivo.query.filter_by(serial_number=serial).first()
        if not disp:
            return jsonify({"error":"Dispositivo no encontrado. Asegúrese de que haya sido detectado por MQTT"}), 404

        if disp.reclamado:
            return jsonify({"error":"El dispositivo ya fue reclamado previamente"}), 409

        # Actualiza SOLO si vienen campos
        if "nombre" in data:       disp.nombre = data["nombre"]
        if "tipo" in data:         disp.tipo = data["tipo"]
        if "modelo" in data:       disp.modelo = data["modelo"]
        if "descripcion" in data:  disp.descripcion = data["descripcion"]
        if "configuracion" in data:
            cfg = dict(disp.configuracion or {})
            cfg.update(data["configuracion"] or {})
            disp.configuracion = cfg

        disp.reclamado = True

        # Captura primitivos ANTES de commit
        disp_id = disp.id

        # Log de acción: reclaim
        _log_action(disp.id, "claimed", {"serial_number": disp.serial_number})

        try:
            db.session.commit() # guarda reclamado y cualquier config inicial
            #Dispara Rule2
            now = datetime.now(timezone.utc)
            dispatch_measure(disp, None, None, ts=now)
        except IntegrityError as e:
            db.session.rollback()
            return jsonify({"error":"Conflicto al reclamar", "detalle": str(e)}), 409
        except SQLAlchemyError as e:
            db.session.rollback()
            # Reintento blando: vuelve a leer y verifica estado
            disp = Dispositivo.query.filter_by(serial_number=serial).first()
            if disp and disp.reclamado:
                return jsonify({"mensaje":"Dispositivo reclamado correctamente", "id": disp.id}), 200
            return jsonify({"error":"Error al reclamar dispositivo", "detalle": str(e)}), 500

        # ✅ No toques disp aquí (evita refrescos implícitos)
        return jsonify({"mensaje":"Dispositivo reclamado correctamente", "id": disp_id}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error":"Error al reclamar dispositivo", "detalle": str(e)}), 500


# Nuevo endpoint para obtener dispositivos no reclamados
@bp.route('/dispositivos/no-reclamados', methods=['GET'])
@jwt_required(optional=True)
def obtener_no_reclamados():
    try:
        dispositivos = Dispositivo.query.filter_by(reclamado=False).order_by(Dispositivo.id.desc()).all()
        return jsonify([_device_full_payload(d) for d in dispositivos])
    except Exception as e:
        return jsonify({"error": "Error al obtener dispositivos no reclamados", "detalle": str(e)}), 500

## Helper
def update_dispositivo_from_payload(dispositivo, data, reclamar=False):
    """
    Actualiza un dispositivo desde un payload recibido.
    - Si reclamar=True, también marca el dispositivo como reclamado.
    """
    if "nombre" in data:
        dispositivo.nombre = data["nombre"]
    if "tipo" in data:
        dispositivo.tipo = data["tipo"]
    if "modelo" in data:
        dispositivo.modelo = data["modelo"]
    if "descripcion" in data:
        dispositivo.descripcion = data["descripcion"]
    if "estado" in data:
        dispositivo.estado = data["estado"]
    if "parametros" in data:
        # ahora lo dejamos para el merge en actualizar_dispositivo
        dispositivo.parametros = data["parametros"]
    if "configuracion" in data:
        # ahora lo dejamos para el merge en actualizar_dispositivo
        dispositivo.configuracion = data["configuracion"]

    if reclamar:
        dispositivo.reclamado = True

    return dispositivo

### OPEN METEO ENDPOINTS
# Estado de inyección en memoria (simple)
_METEO_INJECT = {"payload": None, "until": 0.0}
_LAST_METEO = {"ts": 0.0, "data": None}  # cache de respuesta (normalizada)

# ---- normalización de claves a canónicas para Rule5 --------------------
def _normalize_current_payload(raw: dict) -> dict:
    """
    Convierte diversas claves (current de Open-Meteo o payload inyectado) a canónicas:
      temperature, rain, precipitation, windspeed, uv_index, humidity, cloud_cover, shortwave_radiation
    Si un campo no existe, no aparece o queda como None.
    """
    if not isinstance(raw, dict):
        return {}

    # Mapeo de nombres reales en 'current' -> canónicos
    # Referencia común de Open-Meteo 'current' (puede variar por modelo):
    # temperature_2m, wind_speed_10m, uv_index, precipitation, rain, relative_humidity_2m, cloud_cover (a veces), shortwave_radiation (raro en current)
    current_to_canon = {
        "temperature_2m": "temperature",
        "wind_speed_10m": "windspeed",
        "uv_index": "uv_index",
        "precipitation": "precipitation",
        "rain": "rain",
        "relative_humidity_2m": "humidity",
        "cloud_cover": "cloud_cover",
        "shortwave_radiation": "shortwave_radiation",
    }

    # Alias de entrada adicionales (por si inyectas nombres distintos)
    alias = {
        "temp": "temperature",
        "temperature": "temperature",
        "wind": "windspeed",
        "wind_speed": "windspeed",
        "windspeed": "windspeed",
        "uv": "uv_index",
        "humidity": "humidity",
        "clouds": "cloud_cover",
        "sw_rad": "shortwave_radiation",
        "light": "shortwave_radiation",
    }

    out = {}

    # 1) Si viene un paquete completo de open-meteo (con 'current'), tómalo
    src = raw.get("current") if isinstance(raw.get("current"), dict) else raw

    # 2) Mapea current->canónico
    for k, v in src.items():
        if k in current_to_canon:
            out[current_to_canon[k]] = v

    # 3) Aplica alias (útil para payloads de inyección o datasets propios)
    for k, v in src.items():
        if k in alias and alias[k] not in out:
            out[alias[k]] = v

    # 4) Asegura presencia de claves canónicas (aunque None si no existen)
    canons = ["temperature", "rain", "precipitation", "windspeed", "uv_index", "humidity", "cloud_cover", "shortwave_radiation"]
    for c in canons:
        out.setdefault(c, None)

    # Preferencias: si hay rain y precipitation ausente, o viceversa, deja ambos si existen
    # (no se pisa uno con otro; la regla puede usar el que quiera)
    return out

def _build_openmeteo_current_params():
    cfg = current_app.config
    lat = cfg.get("OPENMETEO_LAT")
    lon = cfg.get("OPENMETEO_LON")

    # canónico -> nombre en 'current'
    canon_to_current = {
        "temperature": "temperature_2m",
        "windspeed": "wind_speed_10m",
        "uv_index": "uv_index",
        "rain": "rain",
        "precipitation": "precipitation",
        "humidity": "relative_humidity_2m",
        "cloud_cover": "cloud_cover",
        "shortwave_radiation": "shortwave_radiation",
    }

    want = []
    for f in current_app.config.get("OPENMETEO_CURRENT_FIELDS", []):
        nm = canon_to_current.get(f.strip())
        if nm and nm not in want:
            want.append(nm)
    if not want:
        want = ["temperature_2m","wind_speed_10m","uv_index","rain","relative_humidity_2m"]

    return {
        "base": "https://api.open-meteo.com/v1/forecast",
        "params": {
            "latitude": lat,
            "longitude": lon,
            "current": ",".join(want)
        }
    }

def _fetch_openmeteo_current():
    """Consulta Open-Meteo (solo 'current'), con timeouts, y devuelve dict CANÓNICO."""
    try:
        setup = _build_openmeteo_current_params()
        tconn = int(current_app.config.get("OPENMETEO_CONNECT_TIMEOUT_S", 3))
        tread = int(current_app.config.get("OPENMETEO_READ_TIMEOUT_S", 5))
        r = requests.get(setup["base"], params=setup["params"], timeout=(tconn, tread))
        r.raise_for_status()
        js = r.json()
        cur = js.get("current", {}) or js.get("current_weather", {}) or {}
        return _normalize_current_payload(cur)
    except Exception as e:
        return {"_error": str(e)}

def _get_meteo_live():
    """Lee Open-Meteo/current con cache + respeta inyección (todo normalizado a canónico)."""
    now = time.time()

    # 1) inyección manda (si activa)
    if _METEO_INJECT["payload"] is not None and now < _METEO_INJECT["until"]:
        norm = _normalize_current_payload(_METEO_INJECT["payload"])
        return {"source": "injected", "at": int(now), "data": norm}

    # 2) cache
    ttl = int(current_app.config.get("OPENMETEO_CACHE_TTL_S", 180))
    if _LAST_METEO["data"] is not None and (now - _LAST_METEO["ts"]) < ttl:
        return {"source": "cache", "at": int(_LAST_METEO["ts"]), "data": _LAST_METEO["data"]}

    # 3) fetch
    data = _fetch_openmeteo_current()
    if data and "_error" not in data:
        _LAST_METEO["ts"] = now
        _LAST_METEO["data"] = data
        return {"source": "open-meteo", "at": int(now), "data": data}

    # error
    if _LAST_METEO["data"] is not None:
        return {"source": "stale-cache", "at": int(_LAST_METEO["ts"]), "data": _LAST_METEO["data"], "error": data.get("_error")}
    return {"source": "error", "at": int(now), "error": data.get("_error"), "data": None}

@bp.route("/meteo", methods=["GET"])
def meteo_now():
    """Devuelve el snapshot meteo actual (normalizado a canónico)."""
    return jsonify(_get_meteo_live())

@bp.route("/meteo/inject", methods=["POST"])
def meteo_inject():
    """
    Inyecta datos actuales (normalizados automáticamente).
    body: { "payload": { ... }, "ttl_sec": 60 }
    Ejemplo payload: { "rain": 8.5, "uv_index": 9 } o estilo 'current' de Open-Meteo.
    Si 'payload' es {} o null -> limpia inyección.
    """
    body = request.get_json() or {}
    payload = body.get("payload")
    ttl = int(body.get("ttl_sec", 60))
    now = time.time()

    if isinstance(payload, dict) and payload:
        _METEO_INJECT["payload"] = payload
        _METEO_INJECT["until"] = now + max(1, ttl)
        return jsonify({"ok": True, "mode": "injected", "until": int(_METEO_INJECT["until"])})
    else:
        _METEO_INJECT["payload"] = None
        _METEO_INJECT["until"] = 0.0
        return jsonify({"ok": True, "mode": "cleared"})

@bp.route('/stream/dispositivos', methods=['GET'])
def stream_dispositivos():
    """
    SSE para cambios de dispositivos:
    - JSON compacto (menos bytes)
    - Heartbeat 'ping' cada ~25s para evitar timeouts
    """
    serial_filter = request.args.get('serial')
    recl_filter   = request.args.get('reclamado')

    def gen():
        q = subscribe()
        try:
            yield "event: hello\ndata: {}\n\n"
            last_ping = time.time()
            while True:
                try:
                    evt = q.get(timeout=5)

                    # ⛔️ Excluir eventos de IA
                    if isinstance(evt, dict) and str(evt.get("event","")).startswith("ai_"):
                        continue

                    # Filtros opcionales
                    if serial_filter and evt.get("serial_number") != serial_filter:
                        continue
                    if recl_filter in ('true','false') and str(evt.get("reclamado")).lower() != recl_filter:
                        continue

                    payload = json.dumps(evt, ensure_ascii=False, separators=(',',':'))
                    # Opcional: usa nombre de evento si viene, o 'device_update' por defecto
                    evname = evt.get("event") or "device_update"
                    yield f"event: {evname}\n"
                    yield f"data: {payload}\n\n"

                except Empty:
                    if time.time() - last_ping > 25:
                        yield "event: ping\ndata: {}\n\n"
                        last_ping = time.time()
        finally:
            unsubscribe(q)

    headers = {
        "Content-Type": "text/event-stream; charset=utf-8",
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
        "Connection": "keep-alive",
    }
    return Response(stream_with_context(gen()), headers=headers)


# =========================================================
# PERFIL / SEGURIDAD
# =========================================================

def _current_user() -> User | None:
    """Obtiene el usuario actual (JWT)."""
    user_id_raw = get_jwt_identity()
    try:
        user_id = int(user_id_raw) if user_id_raw is not None else None
    except (TypeError, ValueError):
        return None
    if user_id is None:
        return None
    return User.query.get(user_id)


def _ensure_avatar_dirs():
    """Asegura que exista static/avatars/."""
    static_folder = current_app.static_folder or "static"
    avatars_dir = os.path.join(static_folder, "avatars")
    os.makedirs(avatars_dir, exist_ok=True)
    return avatars_dir


def _public_avatar_url(rel_path: str) -> str:
    """
    Construye URL pública absoluta hacia /static/<rel_path>.
    (Útil para recursos accesibles por HTTP; NO lo uses para imágenes en correo).
    """
    base = request.url_root.rstrip("/")
    rel = rel_path.lstrip("/")
    return f"{base}/static/{rel}"


def _get_or_create_profile_extra(u: User) -> UserProfileExtra:
    """Obtiene o crea UserProfileExtra para el usuario."""
    extra = UserProfileExtra.query.filter_by(user_id=u.id).first()
    if not extra:
        extra = UserProfileExtra(user_id=u.id, avatar_path="")
        # Si el modelo ya tiene columna 'theme', inicializa en 'dark' (no rompe si no existe)
        try:
            if getattr(extra, "theme", None) is None:
                setattr(extra, "theme", "dark")
        except Exception:
            pass
        db.session.add(extra)
        db.session.commit()
    return extra


def _effective_theme(extra: UserProfileExtra | None) -> str:
    """Devuelve 'dark' o 'light' desde el perfil extra (por defecto 'dark')."""
    if not extra:
        return "dark"
    # getattr evita crashear si aún no existe la columna 'theme'
    return (getattr(extra, "theme", None) or "dark").lower()


def _user_profile_payload(u: User) -> dict:
    """Devuelve el payload del perfil con avatarUrl y theme (camelCase)."""
    extra = UserProfileExtra.query.filter_by(user_id=u.id).first()
    avatar_path = ""
    if extra and extra.avatar_path:
        avatar_path = _public_avatar_url(extra.avatar_path)
    theme = _effective_theme(extra)
    return {
        "id": u.id,
        "email": u.email,
        "nombre": u.nombre or "",
        "avatarUrl": avatar_path,
        "theme": theme
    }


def _generate_code(n_digits: int = 6) -> str:
    """Código aleatorio numérico con n dígitos (default 6)."""
    upper = 10 ** n_digits
    num = randbelow(upper)
    return str(num).zfill(n_digits)

# ------------------ Email helpers (bonitos) ------------------

def _smtp_config():
    cfg = current_app.config
    return {
        "host": cfg.get("SMTP_HOST", "smtp.gmail.com"),
        "port": int(cfg.get("SMTP_PORT", 465)),
        "user": cfg.get("SMTP_USER"),
        "passwd": cfg.get("SMTP_PASS"),
        "from_name": cfg.get("SMTP_FROM_NAME", "Smarthome"),
        "use_ssl": bool(cfg.get("SMTP_USE_SSL", True)),
        "use_tls": bool(cfg.get("SMTP_USE_TLS", False)),
    }

def _email_logo_abspath() -> str:
    """
    Ruta absoluta del logo (rayito azul) dentro del proyecto.
    Ubicación: app/static/Email/bolt-blue.png
    """
    static_folder = current_app.static_folder or os.path.join(os.path.dirname(__file__), "static")
    return os.path.join(static_folder, "Email", "bolt-blue.png")

def _send_email_html(to_email: str, subject: str, html: str, text: str = "", inline_images: list[dict] | None = None):
    """
    Envía correo HTML. Si 'inline_images' viene con dicts {cid, path, maintype, subtype, filename},
    las incrusta como multipart/related y se referencian con <cid> en el HTML (src="cid:cid")."
    """
    smtp = _smtp_config()
    from_addr = smtp["user"]
    if not from_addr:
        raise RuntimeError("SMTP_USER no configurado")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f'{smtp["from_name"]} <{from_addr}>'
    msg["To"] = to_email

    if not text:
        text = "Este correo contiene contenido HTML. Si no lo ves, habilita HTML o usa un cliente compatible."
    msg.set_content(text)

    # Alternativa HTML
    msg.add_alternative(html, subtype="html")

    # Si hay imágenes inline, añadirlas como 'related' al HTML part
    if inline_images:
        html_part = msg.get_payload()[-1]  # última parte (text/html)
        for img in inline_images:
            try:
                with open(img["path"], "rb") as f:
                    data = f.read()
                html_part.add_related(
                    data,
                    maintype=img.get("maintype", "image"),
                    subtype=img.get("subtype", "png"),
                    cid=f"<{img['cid']}>",
                    filename=img.get("filename", os.path.basename(img["path"]))
                )
            except Exception as e:
                current_app.logger.warning(f"[email] No se pudo adjuntar imagen inline {img['path']}: {e}")

    # Envío por SSL/TLS
    if smtp["use_ssl"]:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp["host"], smtp["port"], context=context) as server:
            server.login(smtp["user"], smtp["passwd"])
            server.send_message(msg)
    else:
        with smtplib.SMTP(smtp["host"], smtp["port"]) as server:
            server.ehlo()
            if smtp["use_tls"]:
                server.starttls()
                server.ehlo()
            server.login(smtp["user"], smtp["passwd"])
            server.send_message(msg)

def _email_theme():
    # Paleta y estilos base (oscuro + azul)
    return {
        "bg": "#0b1220",
        "card": "#111827",
        "border": "#1f2937",
        "text_primary": "#e5e7eb",
        "text_secondary": "#9ca3af",
        "primary": "#3B82F6",   # azul
        "code_bg": "#0ea5ff22", # celeste translúcido
    }

def _render_code_email_html(app_name: str, code: str, user_email: str, logo_cid: str) -> tuple[str, str, str]:
    """
    Devuelve (subject, html, text) para el correo de código de verificación.
    El logo se referencia como cid:logo_cid para garantizar que aparezca en clientes como Gmail.
    """
    th = _email_theme()
    subject = f"{app_name} • Código de verificación"
    # Texto plano (fallback)
    text = (
        f"{app_name}\n\n"
        f"Tu código de verificación es: {code}\n"
        "Este código vence en 5 minutos. Si no solicitaste este cambio, ignora este mensaje."
    )

    # ✅ Ajuste de centrado y tamaño del rayo:
    #   - Evitamos flexbox (algunos clientes lo ignoran) y usamos line-height + text-align (compatibilidad alta).
    #   - Imagen más grande (44x44) y centrada vertical/horizontal dentro de un círculo de 64x64.
    html = f"""
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="color-scheme" content="dark light">
  <meta name="supported-color-schemes" content="dark light">
  <title>{subject}</title>
</head>
<body style="margin:0;background:{th['bg']};font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Helvetica,Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:{th['bg']};padding:24px 0;">
    <tr>
      <td align="center">
        <table role="presentation" width="560" cellpadding="0" cellspacing="0" style="background:{th['card']};border:1px solid {th['border']};border-radius:14px;overflow:hidden;">
          <tr>
            <td style="padding:24px 24px 12px 24px;" align="center">
              <div style="width:64px;height:64px;border-radius:999px;background:linear-gradient(180deg,{th['primary']}33,transparent);text-align:center;line-height:64px;">
                <img src="cid:{logo_cid}" width="44" height="44" alt="⚡" style="display:inline-block;vertical-align:middle;border:0;outline:none;">
              </div>
              <h1 style="margin:16px 0 6px 0;color:{th['text_primary']};font-size:20px;line-height:28px;">Código de verificación</h1>
              <p style="margin:0;color:{th['text_secondary']};font-size:14px;line-height:22px;">
                Enviamos este código a <strong style="color:{th['text_primary']};">{user_email}</strong>
              </p>
            </td>
          </tr>

          <tr>
            <td align="center" style="padding:8px 24px 24px 24px;">
              <div style="display:inline-block;padding:14px 18px;border-radius:12px;background:{th['code_bg']};border:1px dashed {th['primary']};">
                <code style="color:{th['primary']};font-weight:700;letter-spacing:8px;font-size:28px;display:inline-block;">
                  {code}
                </code>
              </div>
              <p style="margin:14px 0 0 0;color:{th['text_secondary']};font-size:13px;">
                Vence en 5 minutos.
              </p>
            </td>
          </tr>

          <tr>
            <td style="padding:0 24px 18px 24px;">
              <div style="height:1px;background:{th['border']};"></div>
            </td>
          </tr>

          <tr>
            <td style="padding:0 24px 24px 24px;color:{th['text_secondary']};font-size:13px;line-height:20px;">
              <p style="margin:0;">
                Si no solicitaste cambiar tu contraseña, puedes ignorar este correo con tranquilidad.
              </p>
            </td>
          </tr>

          <tr>
            <td align="center" style="padding:0 24px 24px 24px;color:{th['text_secondary']};font-size:12px;">
              <span style="opacity:.8">© {datetime.utcnow().year} {app_name}. Todos los derechos reservados.</span>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>
    """.strip()

    return subject, html, text

def _send_password_code_email(to_email: str, code: str):
    """
    Envía el email de verificación con el rayo azul incrustado inline (cid),
    para que siempre se vea (sin depender de URL pública).
    """
    cfg = current_app.config
    app_name = cfg.get("SMTP_FROM_NAME", "Smarthome")

    # Generar un CID estable para el logo (sin los <>)
    logo_cid = make_msgid(domain="smarthome.local")[1:-1]
    subject, html, text = _render_code_email_html(app_name, code, to_email, logo_cid)

    # Ruta absoluta al png del rayo
    logo_path = _email_logo_abspath()

    inline = []
    if os.path.isfile(logo_path):
        inline.append({
            "cid": logo_cid,
            "path": logo_path,
            "maintype": "image",
            "subtype": "png",
            "filename": "bolt-blue.png",
        })
    else:
        current_app.logger.warning(f"[email] Logo no encontrado en {logo_path}. Enviando sin imagen.")

    _send_email_html(to_email, subject, html, text, inline_images=inline)

# =========================================================
# AUTH
# =========================================================

@bp.route("/auth/register", methods=["POST"])
def auth_register():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()
    nombre = (data.get("nombre") or "").strip()

    if not email or not password:
        return jsonify({"error": "Email y password son obligatorios"}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email ya registrado"}), 409

    u = User(email=email, nombre=nombre)
    u.set_password(password)
    db.session.add(u)
    db.session.commit()

    token = create_access_token(identity=str(u.id))  # subject como string
    return jsonify({"access_token": token, "user": {"id": u.id, "email": u.email, "nombre": u.nombre}}), 201


@bp.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()

    u = User.query.filter_by(email=email).first()
    if not u or not u.check_password(password):
        return jsonify({"error": "Credenciales inválidas"}), 401

    token = create_access_token(identity=str(u.id))  # subject como string
    return jsonify({"access_token": token, "user": {"id": u.id, "email": u.email, "nombre": u.nombre}})

# --------- Registro con verificación por correo ---------

@bp.route("/auth/register/start", methods=["POST"])
def auth_register_start():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"error": "Email es obligatorio"}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email ya registrado"}), 409

    tmp_password = secrets.token_hex(16)
    u = User(email=email, nombre="")
    u.set_password(tmp_password)
    db.session.add(u)
    db.session.commit()

    SecurityCode.query.filter_by(user_id=u.id, purpose="register_verify", consumed=False).delete()

    code = _generate_code(6)
    expires = datetime.utcnow() + timedelta(minutes=5)
    sc = SecurityCode(user_id=u.id, purpose="register_verify", code=code, expires_at=expires)
    db.session.add(sc)
    db.session.commit()

    try:
        _send_password_code_email(u.email, code)
        current_app.logger.info(f"[register] Código de verificación enviado a {u.email}")
    except Exception as e:
        current_app.logger.error(f"[SMTP] Error enviando email de registro a {u.email}: {e}")
        try:
            SecurityCode.query.filter_by(user_id=u.id, purpose="register_verify").delete()
            db.session.delete(u)
            db.session.commit()
        except Exception:
            pass
        return jsonify({"error": "No se pudo enviar el código por email"}), 500

    return jsonify({"ok": True, "message": "Código enviado al correo"})

@bp.route("/auth/register/verify", methods=["POST"])
def auth_register_verify():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    code = (data.get("code") or "").strip()
    nombre = (data.get("nombre") or "").strip()
    password = (data.get("password") or "").strip()

    if not email or not code or not password:
        return jsonify({"error": "Faltan campos"}), 400

    u = User.query.filter_by(email=email).first()
    if not u:
        return jsonify({"error": "Email no encontrado o no inició registro"}), 404

    sc = SecurityCode.query.filter_by(user_id=u.id, purpose="register_verify", code=code, consumed=False).first()
    if not sc:
        return jsonify({"error": "Código inválido"}), 400

    if datetime.utcnow() > sc.expires_at:
        return jsonify({"error": "Código expirado"}), 400

    sc.consumed = True
    if nombre:
        u.nombre = nombre
    u.set_password(password)
    db.session.commit()

    return jsonify({"ok": True, "message": "Cuenta creada correctamente"})

# --------- Recuperación de contraseña (sin sesión) ---------

@bp.route("/auth/forgot/start", methods=["POST"])
def auth_forgot_start():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"error": "Email es obligatorio"}), 400

    u = User.query.filter_by(email=email).first()
    if not u:
        return jsonify({"error": "Email no encontrado"}), 404

    SecurityCode.query.filter_by(user_id=u.id, purpose="forgot_password", consumed=False).delete()

    code = _generate_code(6)
    expires = datetime.utcnow() + timedelta(minutes=5)
    sc = SecurityCode(user_id=u.id, purpose="forgot_password", code=code, expires_at=expires)
    db.session.add(sc)
    db.session.commit()

    try:
        _send_password_code_email(u.email, code)
        current_app.logger.info(f"[forgot] Código enviado a {u.email}")
    except Exception as e:
        current_app.logger.error(f"[SMTP] Error enviando email a {u.email}: {e}")
        try:
            db.session.delete(sc)
            db.session.commit()
        except Exception:
            pass
        return jsonify({"error": "No se pudo enviar el código por email"}), 500

    return jsonify({"ok": True, "message": "Código enviado al correo"})

@bp.route("/auth/forgot/verify", methods=["POST"])
def auth_forgot_verify():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    code = (data.get("code") or "").strip()
    new_password = (data.get("new_password") or "").strip()

    if not email or not code or not new_password:
        return jsonify({"error": "Faltan campos"}), 400

    u = User.query.filter_by(email=email).first()
    if not u:
        return jsonify({"error": "Email no encontrado"}), 404

    sc = SecurityCode.query.filter_by(user_id=u.id, purpose="forgot_password", code=code, consumed=False).first()
    if not sc:
        return jsonify({"error": "Código inválido"}), 400

    if datetime.utcnow() > sc.expires_at:
        return jsonify({"error": "Código expirado"}), 400

    sc.consumed = True
    u.set_password(new_password)
    db.session.commit()

    return jsonify({"ok": True, "message": "Contraseña actualizada"})

# =========================================================
# PERFIL / SEGURIDAD
# =========================================================

@bp.route("/users/me", methods=["GET"])
@jwt_required()
def get_me():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401
    return jsonify(_user_profile_payload(u))


@bp.route("/users/me", methods=["PUT"])
@jwt_required()
def update_me():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    data = request.get_json() or {}
    nombre = (data.get("nombre") or "").strip()

    if nombre:
        u.nombre = nombre

    db.session.commit()
    return jsonify(_user_profile_payload(u))

# --------- Avatar ---------

@bp.route("/users/me/avatar", methods=["POST"])
@jwt_required()
def upload_avatar():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    if "avatar" not in request.files:
        return jsonify({"error": "Falta archivo 'avatar'"}), 400

    file = request.files["avatar"]
    if file.filename == "":
        return jsonify({"error": "Archivo vacío"}), 400

    _ensure_avatar_dirs()
    static_folder = current_app.static_folder or "static"
    base_name = secure_filename(file.filename)
    _name_part, ext = os.path.splitext(base_name)
    if not ext:
        ext = ".jpg"
    fname = f"u{u.id}_{int(time.time())}{ext}"
    rel_path = os.path.join("avatars", fname)
    abs_path = os.path.join(static_folder, rel_path)

    file.save(abs_path)

    extra = UserProfileExtra.query.filter_by(user_id=u.id).first()
    if not extra:
        extra = UserProfileExtra(user_id=u.id, avatar_path=rel_path.replace("\\", "/"))
        # inicializar tema si existe columna
        try:
            if getattr(extra, "theme", None) is None:
                setattr(extra, "theme", "dark")
        except Exception:
            pass
        db.session.add(extra)
    else:
        try:
            if extra.avatar_path:
                old_abs = os.path.join(static_folder, extra.avatar_path)
                if os.path.isfile(old_abs):
                    os.remove(old_abs)
        except Exception:
            pass
        extra.avatar_path = rel_path.replace("\\", "/")

    db.session.commit()

    return jsonify({
        "avatarUrl": _public_avatar_url(extra.avatar_path),
        "theme": _effective_theme(extra)
    })


@bp.route("/users/me/avatar", methods=["DELETE"])
@jwt_required()
def delete_avatar():
    """
    Limpia el avatar del usuario y (si existe) elimina el archivo en disco.
    Devuelve el perfil completo para parseo consistente en el cliente.
    """
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    static_folder = current_app.static_folder or "static"
    extra = UserProfileExtra.query.filter_by(user_id=u.id).first()

    if extra:
        try:
            if extra.avatar_path:
                abs_path = os.path.join(static_folder, extra.avatar_path)
                if os.path.isfile(abs_path):
                    os.remove(abs_path)
        except Exception:
            pass
        extra.avatar_path = ""
        db.session.commit()
    else:
        extra = UserProfileExtra(user_id=u.id, avatar_path="")
        db.session.add(extra)
        db.session.commit()

    return jsonify(_user_profile_payload(u))

# --------- Preferencia de Tema (NUEVO) ---------

@bp.route("/users/me/theme", methods=["GET"])
@jwt_required()
def get_my_theme():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401
    extra = UserProfileExtra.query.filter_by(user_id=u.id).first()
    return jsonify({"theme": _effective_theme(extra)})

@bp.route("/users/me/theme", methods=["PUT", "PATCH"])
@jwt_required()
def set_my_theme():
    """
    Body: {"theme":"light"} | {"theme":"dark"}
    """
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    data = request.get_json() or {}
    theme = (data.get("theme") or "").strip().lower()
    if theme not in ("light", "dark"):
        return jsonify({"error": "Valor de theme inválido (usa 'light' o 'dark')"}), 400

    extra = _get_or_create_profile_extra(u)

    # Si la columna no existe aún, setattr no crashea pero no persistirá:
    try:
        setattr(extra, "theme", theme)
    except Exception:
        pass

    try:
        db.session.commit()
    except Exception as e:
        current_app.logger.error(f"[theme] Error guardando tema: {e}")
        db.session.rollback()
        return jsonify({"error": "No se pudo guardar el tema"}), 500

    return jsonify({"ok": True, "theme": theme})

# Alias opcional /users/me/preferences  (GET/PUT) -> usa misma lógica
@bp.route("/users/me/preferences", methods=["GET"])
@jwt_required()
def get_preferences():
    return get_my_theme()

@bp.route("/users/me/preferences", methods=["PUT", "PATCH"])
@jwt_required()
def set_preferences():
    return set_my_theme()

# --------- Cambio de contraseña / email ---------

@bp.route("/users/me/change-password/start", methods=["POST"])
@jwt_required()
def start_change_password():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    SecurityCode.query.filter_by(user_id=u.id, purpose="change_password", consumed=False).delete()

    code = _generate_code(6)
    expires = datetime.utcnow() + timedelta(minutes=5)
    sc = SecurityCode(user_id=u.id, purpose="change_password", code=code, expires_at=expires)
    db.session.add(sc)
    db.session.commit()

    try:
        _send_password_code_email(u.email, code)
        current_app.logger.info(f"[change-password] Código enviado a {u.email}")
    except Exception as e:
        current_app.logger.error(f"[SMTP] Error enviando email a {u.email}: {e}")
        try:
            db.session.delete(sc)
            db.session.commit()
        except Exception:
            pass
        return jsonify({"error": "No se pudo enviar el código por email"}), 500

    return jsonify({"ok": True, "message": "Código enviado al correo"})


@bp.route("/users/me/change-password/verify", methods=["POST"])
@jwt_required()
def verify_change_password():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    data = request.get_json() or {}
    code = (data.get("code") or "").strip()
    new_password = (data.get("new_password") or "").strip()

    if not code or not new_password:
        return jsonify({"error": "Faltan campos"}), 400

    sc = SecurityCode.query.filter_by(user_id=u.id, purpose="change_password", code=code, consumed=False).first()
    if not sc:
        return jsonify({"error": "Código inválido"}), 400

    if datetime.utcnow() > sc.expires_at:
        return jsonify({"error": "Código expirado"}), 400

    sc.consumed = True
    u.set_password(new_password)
    db.session.commit()

    return jsonify({"ok": True, "message": "Contraseña actualizada"})


@bp.route("/users/me/change-email", methods=["POST"])
@jwt_required()
def change_email():
    u = _current_user()
    if not u:
        return jsonify({"error": "No autenticado"}), 401

    data = request.get_json() or {}
    new_email = (data.get("new_email") or "").strip().lower()
    current_password = (data.get("current_password") or "").strip()

    if not new_email or not current_password:
        return jsonify({"error": "Faltan campos"}), 400

    if not u.check_password(current_password):
        return jsonify({"error": "Contraseña actual inválida"}), 401

    exists = User.query.filter(User.email == new_email, User.id != u.id).first()
    if exists:
        return jsonify({"error": "Email ya en uso"}), 409

    u.email = new_email
    db.session.commit()

    return jsonify({"ok": True, "message": "Email actualizado"})

# =========================================================
# HABITACIONES
# =========================================================

@bp.route("/habitaciones", methods=["GET"])
@jwt_required(optional=True)
def listar_habitaciones():
    rooms = Habitacion.query.order_by(Habitacion.id.desc()).all()
    return jsonify([_room_to_dict(r) for r in rooms])


@bp.route("/habitaciones", methods=["POST"])
@jwt_required()
def crear_habitacion():
    data = request.get_json() or {}
    nombre = (data.get("nombre") or "").strip()
    icon = (data.get("icon") or data.get("icono") or data.get("icon_name") or "").strip()

    if not nombre:
        return jsonify({"error": "Nombre es obligatorio"}), 400

    user_id_raw = get_jwt_identity()
    try:
        user_id = int(user_id_raw) if user_id_raw is not None else None
    except (TypeError, ValueError):
        return jsonify({"error": "Token inválido"}), 401

    room = Habitacion(nombre=nombre, user_id=user_id)
    _set_room_icon_if_present(room, icon)

    db.session.add(room)
    db.session.commit()

    return jsonify(_room_to_dict(room)), 201


@bp.route("/habitaciones/<int:hid>", methods=["DELETE"])
@jwt_required()
def borrar_habitacion(hid):
    room = Habitacion.query.get_or_404(hid)
    Dispositivo.query.filter_by(habitacion_id=room.id).update({"habitacion_id": None})
    db.session.delete(room)
    db.session.commit()
    sse_publish({"event": "room_deleted", "data": {"id": hid}})
    return jsonify({"ok": True})


# ---- Variantes de actualización de habitación

def _actualizar_habitacion_core(hid: int, data: dict):
    room = Habitacion.query.get_or_404(hid)

    nombre = data.get("nombre") or data.get("name")
    if isinstance(nombre, str) and nombre.strip():
        room.nombre = nombre.strip()

    icon = data.get("icon") or data.get("icono") or data.get("icon_name")
    if isinstance(icon, str) and icon.strip():
        _set_room_icon_if_present(room, icon.strip())

    db.session.commit()
    sse_publish({"event": "room_updated", "data": _room_to_dict(room)})
    return jsonify(_room_to_dict(room))


@bp.route("/habitaciones/<int:hid>", methods=["PUT", "PATCH"])
@jwt_required()
def actualizar_habitacion(hid):
    data = request.get_json() or {}
    return _actualizar_habitacion_core(hid, data)


@bp.route("/habitaciones/<int:hid>/update", methods=["POST"])
@jwt_required()
def actualizar_habitacion_post(hid):
    data = request.get_json() or {}
    return _actualizar_habitacion_core(hid, data)


@bp.route("/habitaciones/<int:hid>/editar", methods=["POST"])
@jwt_required()
def editar_habitacion_post(hid):
    data = request.get_json() or {}
    return _actualizar_habitacion_core(hid, data)


@bp.route("/habitaciones/editar/<int:hid>", methods=["POST"])
@jwt_required()
def editar_habitacion_alt(hid):
    data = request.get_json() or {}
    return _actualizar_habitacion_core(hid, data)


@bp.route("/habitaciones/actualizar/<int:hid>", methods=["POST"])
@jwt_required()
def actualizar_habitacion_alt(hid):
    data = request.get_json() or {}
    return _actualizar_habitacion_core(hid, data)


# ---- Dispositivos por habitación

@bp.route("/habitaciones/<int:hid>/dispositivos", methods=["GET"])
@jwt_required(optional=True)
def dispositivos_de_habitacion(hid):
    Habitacion.query.get_or_404(hid)
    devs = Dispositivo.query.filter_by(habitacion_id=hid).order_by(Dispositivo.id.desc()).all()
    return jsonify([_device_full_payload(d) for d in devs])


@bp.route("/habitaciones/<int:hid>/agregar_dispositivo", methods=["POST"])
@jwt_required()
def agregar_dispositivo_a_habitacion(hid):
    data = request.get_json() or {}
    device_id = data.get("dispositivo_id") or data.get("deviceId") or data.get("device_id")
    if device_id is None:
        return jsonify({"error": "dispositivo_id es obligatorio"}), 400

    room = Habitacion.query.get_or_404(hid)
    dev = Dispositivo.query.get_or_404(int(device_id))

    _dev_set_habitacion_id(dev, room.id)

    # Log de acción
    _log_action(dev.id, "assigned_to_room", {"habitacion_id": room.id, "habitacion": room.nombre})

    db.session.commit()

    sse_publish({"event": "device_moved", "data": {"device_id": _dev_get_id(dev), "to_room": room.id}})
    return jsonify({"ok": True})


@bp.route("/habitaciones/<int:hid>/quitar_dispositivo", methods=["POST"])
@jwt_required()
def quitar_dispositivo_de_habitacion(hid):
    data = request.get_json() or {}
    device_id = data.get("dispositivo_id") or data.get("deviceId") or data.get("device_id")
    if device_id is None:
        return jsonify({"error": "dispositivo_id es obligatorio"}), 400

    Habitacion.query.get_or_404(hid)
    dev = Dispositivo.query.get_or_404(int(device_id))

    _dev_set_habitacion_id(dev, None)

    # Log de acción
    _log_action(dev.id, "removed_from_room", {"habitacion_id": hid})

    db.session.commit()

    sse_publish({"event": "device_moved", "data": {"device_id": _dev_get_id(dev), "to_room": None}})
    return jsonify({"ok": True})

# =========================================================
# DISPOSITIVOS (helpers)
# =========================================================

def _query_all_devices():
    return Dispositivo.query.order_by(Dispositivo.id.desc()).all()


def _query_unclaimed_devices():
    """Devuelve Query si hay columna (reclamado/claimed); si no, filtra en memoria."""
    if hasattr(Dispositivo, "reclamado"):
        return Dispositivo.query.filter_by(reclamado=False).order_by(Dispositivo.id.desc()).all()
    if hasattr(Dispositivo, "claimed"):
        return Dispositivo.query.filter_by(claimed=False).order_by(Dispositivo.id.desc()).all()
    return [d for d in _query_all_devices() if not _dev_get_reclamado(d)]

# =========================================================
# NUEVO: EXPORTACIÓN A EXCEL POR DISPOSITIVO
# =========================================================

def _auto_fit(ws):
    """Auto-ajusta (simple) el ancho de columnas según contenido."""
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column if hasattr(column_cells[0], "column") else column_cells[0].column_letter
        for cell in column_cells:
            try:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > length:
                    length = len(v)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(max(length + 2, 12), 60)

def _build_device_excel(dispositivo: Dispositivo) -> bytes:
    wb = Workbook()

    # --- Hoja 1: Resumen ---
    ws1 = wb.active
    ws1.title = "Resumen"

    cfg = dispositivo.configuracion or {}
    params = dispositivo.parametros or {}
    room_name = None
    try:
        if dispositivo.habitacion_id and dispositivo.habitacion:
            room_name = dispositivo.habitacion.nombre
    except Exception:
        room_name = None

    resumen_rows = [
        ("ID", dispositivo.id),
        ("Serial", dispositivo.serial_number),
        ("Nombre", dispositivo.nombre),
        ("Tipo", dispositivo.tipo),
        ("Modelo", dispositivo.modelo or ""),
        ("Descripción", dispositivo.descripcion or ""),
        ("Estado actual", dispositivo.estado or ""),
        ("Habitación", room_name or "-"),
        ("Reclamado", "Sí" if dispositivo.reclamado else "No"),
        ("Capability", infer_capability(infer_kind(dispositivo.serial_number or "", cfg), cfg)),
        ("Kind", infer_kind(dispositivo.serial_number or "", cfg)),
        ("Configuración (JSON)", json.dumps(cfg, ensure_ascii=False)),
        ("Parámetros (JSON)", json.dumps(params, ensure_ascii=False)),
        ("Generado", datetime.utcnow().isoformat() + "Z"),
    ]
    ws1.append(["Campo", "Valor"])
    for k, v in resumen_rows:
        ws1.append([k, v])
    _auto_fit(ws1)

    # --- Hoja 2: Historial (Acciones + Estados) ---
    ws2 = wb.create_sheet("Historial")

    # Acciones
    acciones = AccionLog.query.filter_by(dispositivo_id=dispositivo.id).order_by(AccionLog.timestamp.asc()).all()
    # Estados
    estados = EstadoLog.query.filter_by(dispositivo_id=dispositivo.id).order_by(EstadoLog.timestamp.asc()).all()

    ws2.append(["Fecha/Hora (UTC)", "Tipo", "Evento/Estado", "Detalle"])
    rows = []
    for a in acciones:
        rows.append((
            a.timestamp.isoformat() + "Z",
            "Acción",
            a.evento,
            json.dumps(a.detalle or {}, ensure_ascii=False)
        ))
    for e in estados:
        rows.append((
            e.timestamp.isoformat() + "Z",
            "Estado",
            e.estado or "",
            json.dumps(e.parametros or {}, ensure_ascii=False)
        ))
    # Ordenar por fecha/hora por si acaso
    rows.sort(key=lambda r: r[0])
    for r in rows:
        ws2.append(list(r))
    _auto_fit(ws2)

    # Serializar a bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

@bp.route("/dispositivos/<int:id>/export", methods=["GET"])
@bp.route("/dispositivos/<int:id>/export_excel", methods=["GET"])  # <-- NUEVO alias para Android
@jwt_required(optional=True)
def exportar_dispositivo_excel(id: int):
    """
    Exporta un Excel (.xlsx) SOLO de ese dispositivo:
      - 'Resumen' con metadatos, configuración y parámetros actuales
      - 'Historial' con Acciones (AccionLog) y Estados (EstadoLog) ordenados por tiempo

    Disponible en dos rutas:
      • /dispositivos/<id>/export           (compatibilidad)
      • /dispositivos/<id>/export_excel     (la que usa el Android)
    """
    try:
        dispositivo = Dispositivo.query.get_or_404(id)
        data = _build_device_excel(dispositivo)
        filename = f"dispositivo_{dispositivo.id}_{dispositivo.serial_number}.xlsx"
        return send_file(
            io.BytesIO(data),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        current_app.logger.error(f"[excel] Error exportando dispositivo {id}: {e}")
        return jsonify({"error": "No se pudo generar el Excel", "detalle": str(e)}), 500